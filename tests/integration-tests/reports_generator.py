# Copyright 2019 Amazon.com, Inc. or its affiliates. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License").
# You may not use this file except in compliance with the License.
# A copy of the License is located at
#
# http://aws.amazon.com/apache2.0/
#
# or in the "LICENSE.txt" file accompanying this file.
# This file is distributed on an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, express or implied.
# See the License for the specific language governing permissions and limitations under the License.
import ast
import datetime
import json
import os
import time
from collections import defaultdict
from typing import List

import boto3
import untangle
from framework.metrics_publisher import Metric, MetricsPublisher
from junitparser import JUnitXml

SECONDS_PER_YEAR = 365 * 24 * 60 * 60


def generate_cw_report(test_results_dir, namespace, aws_region, timestamp_day_start=False, start_timestamp=None):
    """
    Publish tests results to CloudWatch
    :param test_results_dir: dir containing the tests outputs.
    :param namespace: namespace for the CW metric.
    :param aws_region: region where to push the metric.
    :param timestamp_day_start: timestamp of the CW metric equal to the start of the current day (midnight).
    :param start_timestamp: timestamp value to use instead of generating one
    """
    test_report_file = os.path.join(test_results_dir, "test_report.xml")
    if not os.path.isfile(test_report_file):
        generate_junitxml_merged_report(test_results_dir)
    report = generate_json_report(test_results_dir=test_results_dir, save_to_file=False)
    metric_pub = MetricsPublisher(region=aws_region)

    if start_timestamp is not None:
        timestamp = datetime.datetime.fromtimestamp(start_timestamp)
    elif timestamp_day_start:
        timestamp = datetime.datetime.combine(datetime.datetime.now(datetime.timezone.utc), datetime.time())
    else:
        timestamp = datetime.datetime.now(datetime.timezone.utc)
    for category, dictionary in report.items():
        if category == "all":
            _put_metrics(metric_pub, namespace, dictionary, [], timestamp)
        else:
            for dimension, metrics in dictionary.items():
                dimensions = [{"Name": category, "Value": dimension}]
                _put_metrics(metric_pub, namespace, metrics, dimensions, timestamp)


def generate_junitxml_merged_report(test_results_dir):
    """
    Merge all junitxml generated reports in a single one.
    :param test_results_dir: output dir containing the junitxml reports to merge.
    """
    merged_xml = JUnitXml()
    for dir, _, files in os.walk(test_results_dir):
        for file in files:
            if file.endswith("results.xml"):
                merged_xml += JUnitXml.fromfile(os.path.join(dir, file))

    merged_xml.write("{0}/test_report.xml".format(test_results_dir), pretty=True)


def generate_json_report(test_results_dir, save_to_file=True):
    """
    Generate a json report containing a summary of the tests results with details
    for each dimension.
    :param test_results_dir: dir containing the tests outputs.
    :param save_to_file:  True to save to file
    :return: a dictionary containing the computed report.
    """
    test_report_file = os.path.join(test_results_dir, "test_report.xml")
    if not os.path.isfile(test_report_file):
        generate_junitxml_merged_report(test_results_dir)

    result_to_label_mapping = {"skipped": "skipped", "failure": "failures", "error": "errors"}
    results = {"all": _empty_results_dict()}
    xml = untangle.parse(test_report_file)
    for testsuite in xml.testsuites.children:
        for testcase in testsuite.children:
            label = "succeeded"
            for key, value in result_to_label_mapping.items():
                if hasattr(testcase, key):
                    label = value
                    break
            results["all"][label] += 1
            results["all"]["total"] += 1

            if hasattr(testcase, "properties"):
                for property in testcase.properties.children:
                    if property["name"] not in ["region", "os", "instance", "scheduler", "feature"]:
                        continue
                    _record_result(results, property["name"], property["value"], label)

    if save_to_file:
        with open("{0}/test_report.json".format(test_results_dir), "w", encoding="utf-8") as out_f:
            out_f.write(json.dumps(results, indent=4))

    return results


def _record_result(results_dict, dimension, dimension_value, label):
    if dimension not in results_dict:
        results_dict[dimension] = {}
    if dimension_value not in results_dict[dimension]:
        results_dict[dimension].update({dimension_value: _empty_results_dict()})
    results_dict[dimension][dimension_value][label] += 1
    results_dict[dimension][dimension_value]["total"] += 1


def _empty_results_dict():
    return {"total": 0, "skipped": 0, "failures": 0, "errors": 0, "succeeded": 0}


def _put_metrics(
    metric_publisher: MetricsPublisher,
    namespace: str,
    metrics: dict[str, int],
    dimensions: List[dict[str, str]],
    timestamp,
):
    # CloudWatch PutMetric API has a TPS of 150. Setting a rate of 60 metrics per second
    put_metric_sleep_interval = 1.0 / 60
    for key, value in metrics.items():
        metric_publisher.publish_metrics_to_cloudwatch(
            namespace,
            [Metric(key, value, "Count", dimensions)],
        )
        time.sleep(put_metric_sleep_interval)

    failures_errors = metrics["failures"] + metrics["errors"]
    failure_rate = float(failures_errors) / metrics["total"] * 100 if metrics["total"] > 0 else 0
    additional_metrics = [
        {"name": "failures_errors", "value": failures_errors, "unit": "Count"},
        {"name": "failure_rate", "value": failure_rate, "unit": "Percent"},
    ]
    for item in additional_metrics:
        metric_publisher.publish_metrics_to_cloudwatch(
            namespace,
            [Metric(item["name"], item["value"], item["unit"], dimensions, timestamp)],
        )
        time.sleep(put_metric_sleep_interval)


def _scan_dynamodb_table(
    table_name, projection_expression, expression_attribute_names, filter_expression, expression_attribute_values
):
    """Scan a DynamoDB table with pagination and return all items."""
    dynamodb_client = boto3.client("dynamodb", region_name="us-east-1")
    all_items = []
    last_evaluated_key = None
    while True:
        scan_params = {
            "TableName": table_name,
            "ProjectionExpression": projection_expression,
            "FilterExpression": filter_expression,
            "ExpressionAttributeNames": expression_attribute_names,
            "ExpressionAttributeValues": expression_attribute_values,
        }

        # Add ExclusiveStartKey if we're not on the first iteration
        if last_evaluated_key:
            scan_params["ExclusiveStartKey"] = last_evaluated_key

        response = dynamodb_client.scan(**scan_params)
        all_items.extend(response.get("Items", []))

        # Check if there are more items to fetch
        last_evaluated_key = response.get("LastEvaluatedKey")
        if not last_evaluated_key:
            break
    return all_items


def generate_launch_time_report(reports_output_dir):
    current_time = int(time.time())
    one_year_ago = current_time - SECONDS_PER_YEAR

    all_items = _scan_dynamodb_table(
        table_name="ParallelCluster-IntegTest-Metadata",
        projection_expression="#status, #avg_launch, #max_launch, #min_launch, #creation_time, #name, #os, #start_time",
        expression_attribute_names={
            "#call_start_time": "call_start_time",
            "#status": "call_status",
            "#avg_launch": "compute_average_launch_time",
            "#max_launch": "compute_max_launch_time",
            "#min_launch": "compute_min_launch_time",
            "#creation_time": "cluster_creation_time",
            "#name": "name",
            "#os": "os",
            "#start_time": "call_start_time",
        },
        filter_expression="#call_start_time >= :one_year_ago",
        expression_attribute_values={":one_year_ago": {"N": str(one_year_ago)}},
    )
    all_items.sort(key=lambda x: x["call_start_time"]["N"], reverse=True)
    for category_name in ["os", "name"]:
        # Analyze the data by os and by name.
        # Therefore, we can see which os is taking long, or which test is taking long.
        category_name_processing = None
        if category_name == "name":
            category_name_processing = _remove_os_from_string
        for statistics_name in [
            "cluster_creation_time",
            "compute_average_launch_time",
            "compute_min_launch_time",
            "compute_max_launch_time",
        ]:
            if statistics_name in ["cluster_creation_time", "compute_average_launch_time"]:
                statistics_processing = _mean
            elif statistics_name in ["compute_max_launch_time"]:
                statistics_processing = max
            else:
                statistics_processing = min
            result = _get_statistics_by_category(
                all_items,
                category_name,
                statistics_name,
                category_name_processing=category_name_processing,
                statistics_processing=statistics_processing,
            )
            create_report(result, [statistics_name, category_name], reports_output_dir)


def generate_performance_report(reports_output_dir):
    current_time = int(time.time())
    one_year_ago = current_time - SECONDS_PER_YEAR

    all_items = _scan_dynamodb_table(
        table_name="ParallelCluster-PerformanceTest-Metadata",
        projection_expression="#instance, #name, #os, #result, #timestamp, #mpi_variation, #num_instances",
        expression_attribute_names={
            "#instance": "instance",
            "#name": "name",
            "#os": "os",
            "#result": "result",
            "#timestamp": "timestamp",
            "#mpi_variation": "mpi_variation",
            "#num_instances": "num_instances",
        },
        filter_expression="#timestamp >= :one_year_ago",
        expression_attribute_values={":one_year_ago": {"N": str(one_year_ago)}},
    )
    all_items.sort(key=lambda x: x["timestamp"]["N"], reverse=True)
    items_by_name = defaultdict(list)
    category_names = [("name", "S"), ("mpi_variation", "S"), ("num_instances", "N"), ("instance", "S")]
    for item in all_items:
        keys = []
        for category_name, value_type in category_names:
            keys += [item.get(category_name, {}).get(value_type, "")]
        key = "_".join(keys)
        items_by_name[key].append(item)
    result = defaultdict(dict)
    all_benchmark_data = {}  # Collect data for summary report
    os_comparison_data = {}  # Collect data for OS comparison report
    for name, items in items_by_name.items():
        result[name] = _get_statistics_from_result(
            items, name, reports_output_dir, all_benchmark_data, os_comparison_data
        )

    # Generate the summary overview report
    generate_performance_summary(all_benchmark_data, os_comparison_data, reports_output_dir)


def generate_performance_summary(all_benchmark_data, os_comparison_data, reports_output_dir):  # noqa C901
    """Generate a summary Excel file showing performance trends across all benchmarks."""
    import pandas as pd
    from openpyxl.styles import PatternFill
    from scipy import stats

    from pcluster.constants import SUPPORTED_OSES

    number_of_os_rotations = [2, 4, 7, 11, 16, 22, 29, 37]
    timeframes = {}
    for num_of_rotation in number_of_os_rotations:
        num_days = len(SUPPORTED_OSES) * num_of_rotation
        timeframes[f"{num_days} days"] = num_days

    summary_data = []
    current_time = time.time()

    for benchmark_key, data in all_benchmark_data.items():
        row = {"Benchmark": benchmark_key}
        timestamps = data["timestamps"]
        values = data["values"]

        if len(timestamps) < 2:
            for tf_name in timeframes:
                row[tf_name] = None
            summary_data.append(row)
            continue

        for tf_name, tf_days in timeframes.items():
            cutoff = current_time - (tf_days * 24 * 60 * 60)
            # Filter data within this timeframe
            mask = [t >= cutoff for t in timestamps]
            tf_timestamps = [t for t, m in zip(timestamps, mask) if m]
            tf_values = [v for v, m in zip(values, mask) if m]

            if len(tf_timestamps) < 2:
                row[tf_name] = None
                continue

            # Check if data covers the beginning of the timeframe (within 30 days of cutoff)
            oldest_data = min(tf_timestamps)
            if oldest_data > cutoff + (30 * 24 * 60 * 60):
                row[tf_name] = None
                continue

            # Linear regression: convert timestamps to days from start
            t0 = min(tf_timestamps)
            x = [(t - t0) / (24 * 60 * 60) for t in tf_timestamps]  # days
            slope, intercept, _, _, _ = stats.linregress(x, tf_values)

            # Calculate percentage change over the timeframe
            start_val = intercept
            end_val = intercept + slope * (max(x) if x else 0)

            if abs(start_val) > 1e-10:
                pct_change = ((end_val - start_val) / abs(start_val)) * 100
                row[tf_name] = pct_change  # Store as number, format later
            else:
                row[tf_name] = None  # Will display as empty

        summary_data.append(row)

    if not summary_data:
        print("No benchmark data available for summary report")
        return

    # Create DataFrame and write to Excel with conditional formatting
    df = pd.DataFrame(summary_data)
    df = df.set_index("Benchmark")

    filename = os.path.join(reports_output_dir, "performance_summary.xlsx")
    print(f"Creating performance summary: {filename}...")

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=True, sheet_name="Trend")
        worksheet = writer.sheets["Trend"]

        # Color coding based on benchmark type:
        # - osu_bibw: higher is better (bandwidth), so positive = good (green), negative = bad (yellow/red)
        # - All other OSU benchmarks: lower is better (latency), so positive = bad (yellow/red), negative = good (green)
        for row in range(2, len(summary_data) + 2):
            benchmark_name = worksheet.cell(row=row, column=1).value
            is_higher_better = "osu_bibw" in benchmark_name if benchmark_name else False

            for col in range(2, len(timeframes) + 2):
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    val = cell.value
                    # Apply percentage number format
                    cell.number_format = "+0.0%;-0.0%;0.0%"
                    cell.value = val / 100  # Convert to decimal for percentage format

                    # Determine if this change is good or bad based on benchmark type
                    is_regression = (val > 0 and not is_higher_better) or (val < 0 and is_higher_better)

                    if is_regression:
                        if abs(val) > 25:
                            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                        elif abs(val) > 10:
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    else:
                        if abs(val) > 25:
                            cell.fill = PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid")
                        elif abs(val) > 10:
                            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        # OS Comparison sheet
        _write_os_comparison_sheet(writer, os_comparison_data)

    print(f"Performance summary saved: {filename}")


def _write_os_comparison_sheet(writer, os_comparison_data):  # noqa: C901
    """Write OS comparison sheet: average of last 30 days per OS, with best/worst coloring and spread."""
    import pandas as pd
    from openpyxl.styles import PatternFill

    current_time = time.time()
    cutoff_30d = current_time - (30 * 24 * 60 * 60)

    # Collect all OS names across all benchmarks
    all_oses = sorted({os_key for bench in os_comparison_data.values() for os_key in bench})

    rows = []
    for benchmark_key in os_comparison_data:
        row = {"Benchmark": benchmark_key}
        os_averages = {}

        for os_key in all_oses:
            os_data = os_comparison_data[benchmark_key].get(os_key)
            if not os_data:
                row[os_key] = None
                continue
            # Filter to last 30 days and compute average
            recent = [v for v, t in zip(os_data["values"], os_data["timestamps"]) if t >= cutoff_30d]
            if recent:
                avg = sum(recent) / len(recent)
                row[os_key] = avg
                os_averages[os_key] = avg
            else:
                row[os_key] = None

        if not os_averages:
            continue

        # Calculate spread
        if len(os_averages) >= 2:
            best = min(os_averages.values())
            worst = max(os_averages.values())
            if abs(best) > 1e-10:
                row["Spread"] = ((worst - best) / abs(best)) * 100
            else:
                row["Spread"] = None
        else:
            row["Spread"] = None

        rows.append(row)

    if not rows:
        return

    # Ensure column order: Benchmark, OS columns sorted, Spread
    columns = ["Benchmark"] + all_oses + ["Spread"]
    df = pd.DataFrame(rows, columns=columns).set_index("Benchmark")
    df.to_excel(writer, index=True, sheet_name="OS Comparison")
    worksheet = writer.sheets["OS Comparison"]

    os_col_start = 2  # Column B (first OS)
    os_col_end = os_col_start + len(all_oses) - 1
    spread_col = os_col_end + 1

    for row_idx in range(2, len(rows) + 2):
        benchmark_name = worksheet.cell(row=row_idx, column=1).value
        is_higher_better = "osu_bibw" in benchmark_name if benchmark_name else False

        # Find best and worst OS values in this row
        os_values = {}
        for col in range(os_col_start, os_col_end + 1):
            cell = worksheet.cell(row=row_idx, column=col)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                os_values[col] = cell.value

        if len(os_values) >= 2:
            if is_higher_better:
                best_col = max(os_values, key=os_values.get)
                worst_col = min(os_values, key=os_values.get)
            else:
                best_col = min(os_values, key=os_values.get)
                worst_col = max(os_values, key=os_values.get)

            # Color best green, worst red
            worksheet.cell(row=row_idx, column=best_col).fill = PatternFill(
                start_color="90EE90", end_color="90EE90", fill_type="solid"
            )
            worksheet.cell(row=row_idx, column=worst_col).fill = PatternFill(
                start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"
            )

        # Color spread column
        spread_cell = worksheet.cell(row=row_idx, column=spread_col)
        if spread_cell.value is not None and isinstance(spread_cell.value, (int, float)):
            spread_val = spread_cell.value
            spread_cell.number_format = "0.0%"
            spread_cell.value = spread_val / 100
            if spread_val > 50:
                spread_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            elif spread_val > 25:
                spread_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def _append_performance_data(target_dict, os_key, performance, timestamp):
    os_time_key = f"{os_key}-time"
    if os_key not in target_dict:
        target_dict[os_key] = []
        target_dict[os_time_key] = []
    target_dict[os_key].append(float(performance))
    target_dict[os_time_key].append(datetime.datetime.fromtimestamp(int(timestamp)).strftime("%Y-%m-%d %H:%M:%S"))


def _collect_summary_data(all_benchmark_data, benchmark_key, os_key, performance, timestamp):
    """Collect raw data for the summary report."""
    full_key = f"{benchmark_key}_{os_key}"
    if full_key not in all_benchmark_data:
        all_benchmark_data[full_key] = {"timestamps": [], "values": []}
    all_benchmark_data[full_key]["timestamps"].append(int(timestamp))
    all_benchmark_data[full_key]["values"].append(float(performance))


def _collect_os_comparison_data(os_comparison_data, benchmark_key, os_key, performance, timestamp):
    """Collect data keyed by benchmark with OS as a sub-dimension."""
    if benchmark_key not in os_comparison_data:
        os_comparison_data[benchmark_key] = {}
    if os_key not in os_comparison_data[benchmark_key]:
        os_comparison_data[benchmark_key][os_key] = {"timestamps": [], "values": []}
    os_comparison_data[benchmark_key][os_key]["timestamps"].append(int(timestamp))
    os_comparison_data[benchmark_key][os_key]["values"].append(float(performance))


def _get_statistics_from_result(  # noqa: C901
    all_items, name, reports_output_dir, all_benchmark_data=None, os_comparison_data=None
):
    result = {}
    result_with_single_layer = {}
    for item in all_items:
        this_result = ast.literal_eval(item["result"]["S"])
        # The result can contain different format because we historical development of the test data collection.
        # 1. The result can be empty or empty list because testing didn't successfully push data in the initial coding.
        # 2. The result can be a list of tuples or a dictionary. For example, the following result shows OSU latency
        # number of corresponding to different packet sizes
        # [('0', '16'), ('1', '16'), ('2', '16'), ... , ('4194304', '1368')]
        # 3. The result can be a single number.
        # For example, OSU barrier test returns a single number without any consideration of packet sizes.
        # The following logic handles all the differences
        if not this_result:
            continue
        os_key = item["os"]["S"]
        timestamp = item["timestamp"]["N"]
        if isinstance(this_result, (dict, list)):
            for key, performance in this_result:
                if key not in result:
                    result[key] = {}
                _append_performance_data(result[key], os_key, performance, timestamp)
                if all_benchmark_data is not None:
                    _collect_summary_data(all_benchmark_data, f"{name}_{key}", os_key, performance, timestamp)
                if os_comparison_data is not None:
                    _collect_os_comparison_data(os_comparison_data, f"{name}_{key}", os_key, performance, timestamp)
        else:
            _append_performance_data(result_with_single_layer, os_key, this_result, timestamp)
            if all_benchmark_data is not None:
                _collect_summary_data(all_benchmark_data, name, os_key, this_result, timestamp)
            if os_comparison_data is not None:
                _collect_os_comparison_data(os_comparison_data, name, os_key, this_result, timestamp)
    for key, node_num_result in result.items():
        create_report(node_num_result, [name, key], reports_output_dir)
    if result_with_single_layer:
        create_report(result_with_single_layer, [name], reports_output_dir)
    return result


def _mean(x):
    return sum(x) / len(x)


def _remove_os_from_string(x):
    from pcluster.constants import SUPPORTED_OSES

    for os_key in SUPPORTED_OSES:
        x = x.replace(os_key, "")
    return x


def _get_statistics_by_category(  # noqa C901
    all_items, category_name, statistics_name, category_name_processing=None, statistics_processing=None
):
    from pcluster.constants import SUPPORTED_OSES

    # This function is used to get "cluster_creation_time", "compute_average_launch_time",
    # "compute_min_launch_time", and "compute_max_launch_time",
    # This function uses a window of the number of operating systems,
    # so that the statistics are more stable when os rotation is in place.
    more_data = True
    latest_time = float(all_items[0]["call_start_time"]["N"])
    window_length = len(SUPPORTED_OSES)
    result = {}
    while more_data:
        more_data = False
        os_cluster_creation_times = {}
        for item in all_items:
            if item["call_status"]["S"] != "passed":
                continue
            if statistics_name not in item:
                continue
            if float(item["call_start_time"]["N"]) < latest_time - (window_length * 24 * 60 * 60):
                more_data = True
                continue
            if float(item["call_start_time"]["N"]) > latest_time:
                continue
            cluster_creation_time = item[statistics_name]["N"]
            if cluster_creation_time == "0":
                continue
            os_key = item[category_name]["S"]
            if category_name_processing:
                os_key = category_name_processing(os_key)
            if os_key not in os_cluster_creation_times:
                os_cluster_creation_times[os_key] = [float(cluster_creation_time)]
            else:
                os_cluster_creation_times[os_key].append(float(cluster_creation_time))
        for os_key, cluster_creation_times in os_cluster_creation_times.items():
            if os_key not in result:
                result[os_key] = []
            os_time_key = f"{os_key}-time"
            if os_time_key not in result:
                result[os_time_key] = []
            result[os_key].insert(0, sum(cluster_creation_times) / len(cluster_creation_times))
            result[os_time_key].insert(0, datetime.datetime.fromtimestamp(latest_time).strftime("%Y-%m-%d"))
        if os_cluster_creation_times:
            more_data = True
        latest_time = latest_time - 24 * 60 * 60

    return result


def plot_statistics(result, name_prefix):
    import matplotlib.pyplot as plt

    plt.figure(figsize=(40, 12))

    # Collect and sort all unique time points
    all_times = set()
    for category, values in result.items():
        if "-time" in category:
            all_times.update(values)
    sorted_times = sorted(all_times)
    time_to_index = {time: i for i, time in enumerate(sorted_times)}

    # Plot each category using numeric x positions
    for category, values in result.items():
        if "-time" in category:
            continue
        x_values = result[f"{category}-time"]
        x_positions = [time_to_index[time] for time in x_values]
        plt.plot(x_positions, values, marker="o", label=category)

    plt.title(f"{name_prefix}")
    plt.xlabel("Latest timestamp")
    plt.ylabel("Value")
    plt.grid(True, linestyle="--", alpha=0.7)
    plt.legend()
    plt.xticks(range(len(sorted_times)), sorted_times, rotation=45)
    plt.tight_layout()
    plt.show()


def create_excel_files(result, name_prefix, reports_output_dir):
    import pandas as pd

    filename = os.path.join(reports_output_dir, f"{name_prefix}_statistics.xlsx")
    print(f"Creating Excel file: {filename}...")

    all_times = sorted({t for k, v in result.items() if "-time" in k for t in v})
    df_data = {
        k: pd.Series(index=result[f"{k}-time"], data=v).groupby(level=0).mean().reindex(all_times)
        for k, v in result.items()
        if "-time" not in k
    }
    df = pd.DataFrame(df_data)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.T.to_excel(writer, index=True)

    print(f"Excel file saved: {filename}")


def _get_launch_time(logs, instance_id):
    for log in logs:
        if instance_id in log["message"]:
            return log["timestamp"]


def create_report(result, labels, reports_output_dir, create_graphs=False, create_excel=True):
    name_prefix = "_".join(map(str, labels))
    if create_excel:
        create_excel_files(result, name_prefix, reports_output_dir)
    if create_graphs:
        plot_statistics(result, name_prefix)
